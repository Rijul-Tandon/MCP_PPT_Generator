from __future__ import annotations

"""Spreadsheet summarization helpers.

This module intentionally does *not* try to become a full spreadsheet ingestion
engine. Its purpose is narrower: collect a compact, repeatable view of the most
useful numeric and textual signals so the planner can incorporate fresh facts on
every run.
"""

import csv
from pathlib import Path
from typing import Iterable


class ExcelContextLibrary:
    """Read spreadsheet-like files and turn them into compact planning context.

    The planner does not want full workbooks. It wants a small, repeatable snapshot
    of the most useful numbers and labels so the LLM can reason over current data
    on every run.
    """

    # These limits keep runtime context short enough for prompts while still
    # capturing enough signal to influence slide planning.
    MAX_SHEETS = 5
    MAX_ROWS = 8
    MAX_SAMPLES = 12

    def summarize_files(self, files: list[Path]) -> dict[str, object]:
        """Return a lightweight summary for each readable file plus warnings.

        We capture warnings instead of failing the whole run because a single bad
        spreadsheet should not block planning if the rest of the context is usable.
        """
        summaries: list[dict[str, object]] = []
        warnings: list[str] = []
        for path in files:
            try:
                summaries.append(self._summarize_file(path))
            except Exception as exc:
                warnings.append(f"Could not read {path.name}: {exc}")
        return {"files": summaries, "warnings": warnings}

    def build_runtime_context_block(self, summary: dict[str, object]) -> str:
        """Turn spreadsheet summaries into plain-text context for each run.

        We do not edit context.txt on disk. Instead, we append this generated block
        to the runtime context passed into the planner/LLM so Excel-derived facts are
        refreshed every time a deck is created.
        """
        files = summary.get("files", [])
        if not files:
            return ""

        lines = [
            "## Runtime Excel Context",
            "The points below were extracted automatically from files in excel_context/ for this run.",
            "Treat them as current project inputs only when they support the business story in context.txt.",
        ]
        for file_summary in files:
            file_name = file_summary.get("fileName", "unknown file")
            lines.append(f"### {file_name}")
            for item in file_summary.get("numericSamples", [])[:6]:
                lines.append(f"- Numeric signal: {item}")
            for item in file_summary.get("textSamples", [])[:4]:
                lines.append(f"- Text signal: {item}")
        return "\n".join(lines)

    def _summarize_file(self, path: Path) -> dict[str, object]:
        """Dispatch by file type.

        Today we support `.xlsx`, `.csv`, and `.tsv`. The output shape is kept
        consistent so the planner does not need to care which file type produced it.
        """
        if path.suffix.lower() == ".xlsx":
            return self._summarize_xlsx(path)
        return self._summarize_delimited(path)

    def _summarize_xlsx(self, path: Path) -> dict[str, object]:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheets: list[dict[str, object]] = []
        numeric_samples: list[str] = []
        text_samples: list[str] = []

        for sheet in workbook.worksheets[: self.MAX_SHEETS]:
            rows = list(sheet.iter_rows(values_only=True, max_row=self.MAX_ROWS))
            scanned_rows = self._compact_rows(rows)
            sheets.append(
                {
                    "name": sheet.title,
                    "rowsScanned": len(scanned_rows),
                    "headerPreview": self._header_preview(scanned_rows),
                }
            )
            # Flatten a small window of values from the sheet and collect short
            # numeric/text samples that the planner can later reference.
            self._collect_samples(
                values=(value for row in scanned_rows for value in row[:8]),
                numeric_prefix=f"{sheet.title}",
                text_prefix=f"{sheet.title}",
                numeric_samples=numeric_samples,
                text_samples=text_samples,
            )

        return {
            "fileName": path.name,
            "type": "xlsx",
            "sheets": sheets,
            "numericSamples": numeric_samples,
            "textSamples": text_samples,
        }

    def _summarize_delimited(self, path: Path) -> dict[str, object]:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = [row for row in reader][: self.MAX_ROWS]

        scanned_rows = self._compact_rows(rows)
        numeric_samples: list[str] = []
        text_samples: list[str] = []
        self._collect_samples(
            values=(value for row in scanned_rows for value in row[:8]),
            numeric_prefix=path.name,
            text_prefix=path.name,
            numeric_samples=numeric_samples,
            text_samples=text_samples,
        )

        return {
            "fileName": path.name,
            "type": path.suffix.lower().lstrip("."),
            "sheets": [{"name": path.stem, "rowsScanned": len(scanned_rows), "headerPreview": self._header_preview(scanned_rows)}],
            "numericSamples": numeric_samples,
            "textSamples": text_samples,
        }

    def _compact_rows(self, rows: list[tuple | list]) -> list[list[object]]:
        """Drop empty cells and empty rows so downstream logic stays simple."""
        compact_rows: list[list[object]] = []
        for row in rows:
            values = [value for value in row if value not in (None, "")]
            if values:
                compact_rows.append(values)
        return compact_rows

    def _header_preview(self, rows: list[list[object]]) -> list[str]:
        if not rows:
            return []
        return [str(value)[:40] for value in rows[0][:6]]

    def _collect_samples(
        self,
        *,
        values: Iterable[object],
        numeric_prefix: str,
        text_prefix: str,
        numeric_samples: list[str],
        text_samples: list[str],
    ) -> None:
        """Collect the first useful numeric/text snippets from a value stream."""
        for value in values:
            if self._append_numeric_sample(value, numeric_prefix, numeric_samples):
                continue
            self._append_text_sample(value, text_prefix, text_samples)

    def _append_numeric_sample(self, value: object, prefix: str, samples: list[str]) -> bool:
        if len(samples) >= self.MAX_SAMPLES:
            return False
        if isinstance(value, (int, float)):
            samples.append(f"{prefix}: {value}")
            return True
        try:
            number = float(str(value).replace(",", ""))
        except ValueError:
            return False
        samples.append(f"{prefix}: {number}")
        return True

    def _append_text_sample(self, value: object, prefix: str, samples: list[str]) -> None:
        """Keep text samples compact so prompts do not fill up with spreadsheet noise."""
        if len(samples) >= self.MAX_SAMPLES:
            return
        compact = " ".join(str(value).split())
        if compact:
            samples.append(f"{prefix}: {compact[:80]}")
